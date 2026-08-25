#!/usr/bin/env python3
"""
Generate the app's data files from the Aye Si Cena matrix spreadsheet.

The spreadsheet is the master. Run this after editing it and the app picks up
every change - dishes, costs, sourcing, seasonal links.

    python3 scripts/import-matrix.py path/to/ayesicenamatrix.xlsx

Writes:
    data/dishes.ts     130 dishes with the full schema
    data/sourcing.ts   the supply lines from the Sourcing tab
    data/flavours.ts   flavour axes, auto-derived (see below) + hand overrides

Flavour axes are NOT in the spreadsheet. They are derived here from category
and keyword, then corrected by FLAVOUR_OVERRIDES. Every derived entry is a
guess about palate; the overrides are where real tasting notes go.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

# Spreadsheet category -> app category id
CATEGORY = {
    "Canape": "canape",
    "Main": "main",
    "Side": "side",
    "Breakfast": "breakfast",
    "Bowl": "bowl",
    "Bakery": "bakery",
    "Dessert": "dessert",
}

# Service format constrains which pricing tiers can carry the dish.
# A drop-off item works anywhere; a plated one only at the plated tier.
TIERS_FOR_FORMAT = {
    "Drop-off": ["scran", "buffet", "plated"],
    "Buffet": ["buffet", "plated"],
    "Plated": ["plated"],
    "Live station": ["buffet", "plated"],
}

FORMAT = {
    "Drop-off": "drop-off",
    "Buffet": "buffet",
    "Plated": "plated",
    "Live station": "live-station",
}

# --- flavour derivation -----------------------------------------------------
# Ordered rules; each matching rule contributes an axis. Capped at three.
FLAVOUR_RULES = [
    ("sweet",   r"chocolate|caramel|toffee|honey|manjar|sugar|jam|marmalade|sponge|shortbread|biscuit|cake|dulce|syrup|chancaca|meringue"),
    ("tart",    r"l[uú]cuma|chirimoya|aguaymanto|maracuy[aá]|lim[oó]n|lime|lemon|citrus|orange|vinegar|pickle|piccalilli|cure[d]?|ceviche|leche de tigre|berry|berries|rhubarb|lingonberry|passion"),
    ("smoky",   r"smok|ahumad|peated|charred|grill|barbec|anticucho|bacon|arbroath"),
    ("spiced",  r"aj[ií]|rocoto|panca|amarillo|curry|masala|tikka|ginger|clove|cinnamon|pepper|mustard|spice|mace|nutmeg|cardamom|hu?acatay|mu[nñ]a"),
    ("rich",    r"cream|butter|cheese|fried|deep-fry|duck|pork belly|chicharr[oó]n|confit|custard|frangipane|pastry|puff|marrow|foie|lamb|beef|oxtail|stovies|gratin"),
    ("fresh",   r"salad|watercress|berros|cucumber|mint|herb|raw|crudo|leaf|leaves|slaw|dill|parsley"),
]

SWEET_CATEGORIES = {"bakery", "dessert"}

# Hand-corrected axes. Add real tasting notes here; they beat the rules above.
FLAVOUR_OVERRIDES: dict[int, list[str]] = {}

# --- allergens --------------------------------------------------------------
# NOT DERIVED HERE ANY MORE. This file used to read allergens out of a regex
# over the dish name, the fusion line and the "key local ingredients" column -
# that is, out of marketing copy written to sell the dish. It disagreed with the
# recipe on 165 of 223 dishes, and of the 103 dishes it offered as gluten-free,
# 50 contained gluten. Its vocabulary was eight words, so celery, mustard,
# sesame, soya, sulphites and lupin could not be declared at all.
#
# lib/dietary.ts reads them off the recipe's actual ingredient lines. Every row
# below is written with `allergens: []` and filled in by the second step:
#
#   python3 scripts/import-matrix.py data/ayesicena-matrix.xlsx
#   node --experimental-strip-types --import ./scripts/ts-alias.mjs \
#        scripts/derive-allergens.mjs
#
# __tests__/dietary.test.ts fails if that second step is skipped, so a stale
# file cannot ship.

# --- equipment --------------------------------------------------------------
# What the dish occupies during service. Drives the kitchen-collision check.
EQUIPMENT_RULES = {
    "oven":    r"bake|baked|roast|oven|tart|\bpie\b|pastry|gratin|sponge|cake|scone|bread|puff|empanada|dough|bridie|sausage roll|pasty|loaf|bun\b|biscuit|shortbread",
    "fryer":   r"fried|fry|deep-fr|crumbed|bonbon|croqueta|tequen|tequeñ|tempura|chips|fritter|churro|doughnut",
    "griddle": r"griddle|plancha|tattie scone|pancake|crumpet|sear|toast",
    "hob":     r"brais|stew|simmer|soup|chowder|sauce|poach|boil|reduc|chupe|skink",
    "cold":    r"cured|chilled|no-bake|fridge|\bcold\b|causa|ceviche|carpaccio|salad|trifle|posset|mousse|cranachan|whipped",
}

# --- seasonal ingredient keywords ------------------------------------------
# Matched against the "Key local ingredients" column to rebuild the dish links.
INGREDIENT_KEYWORDS = {
    "lucuma":        r"l[uú]cuma",
    "chirimoya":     r"chirimoya",
    "aguaymanto":    r"aguaymanto",
    "maracuya":      r"maracuy[aá]",
    "fresa":         r"fresa|strawberr",
    "asparagus":     r"esp[aá]rrago|asparagus",
    "papa-nativa":   r"papa[s]? nativa|native potato|papa amarilla|yellow potato",
    "choclo":        r"choclo",
    "rocoto":        r"rocoto",
    "muna":          r"mu[ñn]a",
    "berros":        r"berros|watercress",
    "langostinos":   r"langostino|prawn|shrimp",
    "trucha-paiche": r"trucha|trout|paiche|corvina|chita",
    "aji-amarillo":  r"aj[ií] amarillo|aj[ií] panca|aj[ií] limo",
    "huacatay":      r"huacatay",
    "camote":        r"camote|sweet potato",
    "cacao":         r"cacao|chocolate",
    "chancaca":      r"chancaca|panela",
    "quinoa-kiwicha": r"quinoa|kiwicha",
    "cochayuyo":     r"cochayuyo",
}


def ts_string(s):
    """Emit a TypeScript double-quoted string."""
    return json.dumps("" if s is None else str(s).strip(), ensure_ascii=False)


def derive_flavours(dish_id, name, fusion, ingredients, category):
    if dish_id in FLAVOUR_OVERRIDES:
        return FLAVOUR_OVERRIDES[dish_id]

    haystack = " ".join(str(x or "") for x in (name, fusion, ingredients)).lower()
    hits = [axis for axis, pattern in FLAVOUR_RULES if re.search(pattern, haystack)]

    if category in SWEET_CATEGORIES and "sweet" not in hits:
        hits.insert(0, "sweet")
    if category not in SWEET_CATEGORIES and "savoury" not in hits:
        hits.append("savoury")
    if not hits:
        hits = ["savoury"]

    # Keep it readable: at most three axes, order preserved.
    seen, out = set(), []
    for h in hits:
        if h not in seen:
            seen.add(h)
            out.append(h)
    return out[:3]


def main():
    if len(sys.argv) < 2:
        sys.exit("usage: import-matrix.py <workbook.xlsx>")
    wb = openpyxl.load_workbook(sys.argv[1], data_only=True)

    rows = list(wb["Matrix"].iter_rows(values_only=True))
    hdr = list(rows[0])
    idx = {h: i for i, h in enumerate(hdr)}
    data = [r for r in rows[1:] if r and r[0] is not None]

    dishes, flavours, ing_links = [], {}, {k: [] for k in INGREDIENT_KEYWORDS}

    for r in data:
        did = int(r[idx["ID"]])
        raw_cat = str(r[idx["Category"]]).strip()
        cat = CATEGORY.get(raw_cat)
        if not cat:
            sys.exit(f"row {did}: unknown category {raw_cat!r}")

        raw_fmt = str(r[idx["Service format"]]).strip()
        if raw_fmt not in TIERS_FOR_FORMAT:
            sys.exit(f"row {did}: unknown service format {raw_fmt!r}")

        cost = float(r[idx["Est. food cost (S/)"]])
        price = float(r[idx["Suggested price (S/)"]])
        if price <= cost:
            sys.exit(f"row {did}: price {price} not above cost {cost}")

        sub_origin = str(r[idx["Sub-origin"]]).strip()
        # "Disputed (...)" in the sheet means the provenance is contested.
        contested = sub_origin.lower().startswith("disputed")

        name = str(r[idx["Dish"]]).strip()
        fusion = str(r[idx["Peruvian fusion"]]).strip()
        key_ings = str(r[idx["Key local ingredients"]]).strip()

        flavours[did] = derive_flavours(did, name, fusion, key_ings, cat)

        text = f"{name} {fusion} {key_ings}".lower()
        # Left empty on purpose; scripts/derive-allergens.mjs fills it from the
        # recipes. See the note above ALLERGEN_RULES' grave.
        allergens: list[str] = []
        equipment = sorted(e for e, pat in EQUIPMENT_RULES.items() if re.search(pat, text))
        if raw_fmt == "Live station" and "griddle" not in equipment:
            equipment.append("griddle")
        if not equipment:
            equipment = ["cold"]

        hay = f"{name} {fusion} {key_ings}".lower()
        for ing_id, pattern in INGREDIENT_KEYWORDS.items():
            if re.search(pattern, hay):
                ing_links[ing_id].append(did)

        dishes.append({
            "id": did,
            "name": name,
            "origin": str(r[idx["Origin dish"]]).strip(),
            "subOrigin": sub_origin,
            "contested": contested,
            "fusion": fusion,
            "category": cat,
            "format": FORMAT[raw_fmt],
            "needsLicence": str(r[idx["Needs liquor licence"]]).strip().lower() == "yes",
            "veg": str(r[idx["Veg"]]).strip().lower() == "yes",
            "keyIngredients": key_ings,
            "source": str(r[idx["Primary Lima source"]]).strip(),
            "cost": round(cost, 2),
            "price": round(price, 2),
            "costVerified": not str(r[idx["Cost verified?"]]).lower().startswith("no"),
            "allergens": allergens,
            "equipment": sorted(equipment),
            "tiers": TIERS_FOR_FORMAT[raw_fmt],
        })

    # ---- data/dishes.ts ----
    lines = [
        'import type { Dish } from "@/lib/dishes";',
        "",
        "/**",
        " * GENERATED FILE - do not edit by hand.",
        " *",
        " * Source: the Aye Si Cena matrix spreadsheet.",
        " * Regenerate with:  python3 scripts/import-matrix.py <workbook.xlsx>",
        " *",
        " * Every cost and price is an ESTIMATE made without a market run. The",
        f" * spreadsheet marks all {len(dishes)} rows Cost verified? = No, and costVerified",
        " * carries that through. Replace with real supplier prices before quoting.",
        " */",
        "export const DISHES: Dish[] = [",
    ]
    for d in dishes:
        lines.append(
            "  { "
            f'id: {d["id"]}, name: {ts_string(d["name"])}, '
            f'origin: {ts_string(d["origin"])}, subOrigin: {ts_string(d["subOrigin"])}, '
            f'contested: {str(d["contested"]).lower()}, '
            f'fusion: {ts_string(d["fusion"])}, '
            f'category: {ts_string(d["category"])}, format: {ts_string(d["format"])}, '
            f'needsLicence: {str(d["needsLicence"]).lower()}, veg: {str(d["veg"]).lower()}, '
            f'keyIngredients: {ts_string(d["keyIngredients"])}, source: {ts_string(d["source"])}, '
            f'cost: {d["cost"]}, price: {d["price"]}, '
            f'costVerified: {str(d["costVerified"]).lower()}, '
            f'allergens: {json.dumps(d["allergens"])}, '
            f'equipment: {json.dumps(d["equipment"])}, '
            f'tiers: {json.dumps(d["tiers"])} }},'
        )
    lines += ["];", ""]
    (ROOT / "data" / "dishes.ts").write_text("\n".join(lines), encoding="utf-8")

    # ---- data/flavours.ts ----
    fl = [
        'import type { Flavour } from "@/lib/dishes";',
        "",
        "/**",
        " * GENERATED FILE - do not edit by hand.",
        " *",
        " * Flavour axes are NOT in the spreadsheet. They are derived from category",
        " * and keyword by scripts/import-matrix.py, which makes them a machine's",
        " * guess at a palate, not a tasting note.",
        " *",
        " * To correct one, add it to FLAVOUR_OVERRIDES in that script and re-run;",
        " * overrides beat the rules and survive re-import.",
        " */",
        "export const FLAVOURS: Record<number, Flavour[]> = {",
    ]
    for did in sorted(flavours):
        fl.append(f"  {did}: {json.dumps(flavours[did])},")
    fl += ["};", ""]
    (ROOT / "data" / "flavours.ts").write_text("\n".join(fl), encoding="utf-8")

    # ---- data/sourcing.ts ----
    src_rows = [r for r in wb["Sourcing"].iter_rows(values_only=True) if r and r[0]]
    sl = [
        'import type { SupplyLine } from "@/lib/dishes";',
        "",
        "/** GENERATED FILE - do not edit by hand. Source: Sourcing tab. */",
        "export const SOURCING: SupplyLine[] = [",
    ]
    for r in src_rows[1:]:
        sl.append(
            "  { "
            f"name: {ts_string(r[0])}, buy: {ts_string(r[1])}, "
            f"why: {ts_string(r[2])}, verify: {ts_string(r[3])} }},"
        )
    sl += ["];", ""]
    (ROOT / "data" / "sourcing.ts").write_text("\n".join(sl), encoding="utf-8")

    print(f"dishes:   {len(dishes)}")
    print(f"sourcing: {len(src_rows) - 1} supply lines")
    empty = [k for k, v in ing_links.items() if not v]
    print(f"ingredient links rebuilt; empty: {empty if empty else 'none'}")
    print(json.dumps({k: len(v) for k, v in ing_links.items()}, indent=0))
    # Emit the links so data/ingredients.ts can be updated to match.
    (ROOT / "data" / ".ingredient-links.json").write_text(
        json.dumps(ing_links, indent=2), encoding="utf-8"
    )


if __name__ == "__main__":
    main()
